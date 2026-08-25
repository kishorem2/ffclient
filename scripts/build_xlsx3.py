import json, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule

ROOT = Path(__file__).resolve().parent.parent
D = json.load(open(Path(os.environ.get("FF_DATA_DIR", ROOT / "data")) / "consensus.json"))
ARIAL = "Arial"
HDR = "1F2430"
thin = Side(style="thin", color="D0D4DC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
POS_FILL = {"QB": "FFF0D5", "RB": "D9F7E6", "WR": "D9EAFB", "TE": "FBDCE9", "K": "E8E2FB", "DST": "D5F2F0"}

FLAG = D["flag"]["spread"]     # single source of truth, computed in agg_three
FP_INFO = D.get("fp_info")
HAS_ADP = any(r.get("delta") is not None for f in ("half", "ppr") for r in D[f]["rows"])
HAS_TREND = any(r.get("trend_add") is not None or r.get("trend_drop") is not None
                for f in ("half", "ppr") for r in D[f]["rows"])

# Columns are dynamic: ECR-ADP and Trend only exist once live data does.
HDRS = (["Rank", "Tier", "PosTier", "Player", "Pos", "PosRank", "Team", "Avg",
         "Winks", "Norris", "FantasyPros", "Range"]
        + (["ECR-ADP"] if HAS_ADP else []) + (["Trend"] if HAS_TREND else [])
        + ["Drafted", "Notes"])
IDX = {h: i for i, h in enumerate(HDRS)}
COL = {h: get_column_letter(i + 1) for i, h in enumerate(HDRS)}
BASE_WIDTH = {"Rank": 7, "Tier": 6, "PosTier": 8, "Player": 24, "Pos": 6, "PosRank": 8,
              "Team": 7, "Avg": 8, "Winks": 8, "Norris": 8, "FantasyPros": 11,
              "Range": 8, "ECR-ADP": 9, "Trend": 10, "Drafted": 10, "Notes": 32}


def trend_str(r):
    a, d = r.get("trend_add"), r.get("trend_drop")
    parts = ([f"+{a:,}"] if a else []) + ([f"-{d:,}"] if d else [])
    return " / ".join(parts) or None


wb = Workbook()
lg = wb.active; lg.title = "Legend"; lg.sheet_view.showGridLines = False

def put(cell, val, size=10, bold=False, color="000000", fill=None, italic=False):
    c = lg[cell]; c.value = val
    c.font = Font(name=ARIAL, size=size, bold=bold, color=color, italic=italic)
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    return c

fp_depth_half = D["half"]["sizes"]["FantasyPros"]
if FP_INFO:
    src_line = (f"Blends Hayden Winks and Josh Norris (Yahoo Sports, Aug 12 2026) with the FantasyPros "
                f"Expert Consensus Ranking ({FP_INFO['last_updated']}, {FP_INFO['experts']} experts, live API).")
    cap_line = (f"Half PPR and Full PPR are separate boards. FantasyPros is pulled in full from the API "
                f"({fp_depth_half} deep); anyone it leaves off is treated as omitted.")
else:
    src_line = ("Blends Hayden Winks and Josh Norris (Yahoo Sports, Aug 12 2026) with the FantasyPros "
                "Expert Consensus Ranking (Aug 25 2026, 111 experts half-PPR / 108 PPR).")
    cap_line = ("Half PPR and Full PPR are separate boards. FantasyPros is capped at its top 300; "
                "anyone it ranks below that is treated as omitted.")

put("A1", "2026 Draft Cheat Sheet — Three-Source Consensus", 15, True)
put("A2", src_line, 9, color="555555")
put("A3", cap_line, 9, color="555555")

put("A5", "WHICH CELLS YOU EDIT", 11, True)
put("A6", "Two columns are yours: Drafted and Notes. The weights and penalties below are also editable. Everything else recalculates — but the rows will not resort themselves, so use the header filters to sort by Avg after changing a weight.", 9)
put("A7", "Drafted", 10, True, fill="FFF3B0"); put("B7", "Pick ME or TAKEN from the dropdown. Rows turn green or grey out.", 9)
put("A8", "Notes", 10, True, fill="FFF3B0"); put("B8", "Your own reads — injury notes, 'do not draft', target round.", 9)

put("A10", "EXAMPLE OF A FILLED ROW", 11, True)
for i, h in enumerate(HDRS):
    c = lg.cell(row=11, column=1+i, value=h)
    c.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HDR); c.border = box
EXAMPLE = {"Rank": 19, "Tier": 7, "PosTier": 4, "Player": "Omarion Hampton", "Pos": "RB",
           "PosRank": 9, "Team": "LAC", "Avg": 19.0, "Winks": 24, "Norris": 15,
           "FantasyPros": 19, "Range": 9, "ECR-ADP": -4, "Trend": "+12,480",
           "Drafted": "ME", "Notes": "volume is there, line is shaky"}
for i, h in enumerate(HDRS):
    c = lg.cell(row=12, column=1+i, value=EXAMPLE[h])
    c.font = Font(name=ARIAL, size=9); c.border = box
    if h in ("Drafted", "Notes"): c.fill = PatternFill("solid", fgColor="FFF3B0")

MEANINGS = [
    ("Tier", "Overall tier. A break means a real drop-off in consensus value."),
    ("PosTier", "Tier within the position — the one to use mid-draft. It tells you how many comparable players remain at that spot."),
    ("Avg", "Weighted average of the three sources. Formula — driven by the weights and penalties below."),
    ("Winks / Norris / FantasyPros", "Each source's own rank. Blank means that source left the player off"
     + (" entirely." if HAS_ADP else " (or ranked them past FantasyPros' top 300).")),
    ("Range", "Highest minus lowest rank across the sources. A big number means they genuinely disagree and the pick is a judgment call."),
]
if HAS_ADP:
    MEANINGS.append(("ECR-ADP", "Expert rank minus room ADP (both FantasyPros, this scoring format). "
                     "Negative (green) = the room lets him fall past his expert rank — value. Positive (red) = the room reaches."))
if HAS_TREND:
    MEANINGS.append(("Trend", "Sleeper trending adds/drops across all leagues, last 48h. Attention, not value — "
                     "it never touches the consensus. Mostly hype and injury reaction pre-draft."))

put("A14", "COLUMN MEANINGS", 11, True)
row = 15
for k, v in MEANINGS:
    put(f"A{row}", k, 10, True); put(f"B{row}", v, 9)
    row += 1

w_hdr = row + 1
put(f"A{w_hdr}", "WEIGHTS", 11, True)
put(f"A{w_hdr+1}", "FantasyPros ECR is itself an average of 100+ experts, so weighting it equal to one analyst would understate the market. It is set to double. Change these to shift the blend — set FantasyPros to 1 to treat all three as equal opinions.", 9, color="555555")
w_first = w_hdr + 2
for r, (label, val) in enumerate([("Winks weight", 1.0), ("Norris weight", 1.0), ("FantasyPros weight", 2.0)], start=w_first):
    put(f"A{r}", label, 9)
    c = put(f"B{r}", val, 9, True, color="0000FF", fill="FFFF00"); c.border = box
WCELL = {"Winks": f"Legend!$B${w_first}", "Norris": f"Legend!$B${w_first+1}",
         "FantasyPros": f"Legend!$B${w_first+2}"}

p_hdr = w_first + 4
put(f"A{p_hdr}", "OMISSION PENALTIES", 11, True)
put(f"A{p_hdr+1}", "When a source leaves a player off entirely, that counts as a rank just past the end of their list rather than being ignored — being omitted from a board is information. These update from the actual list sizes on every rebuild.", 9, color="555555")
p_first = p_hdr + 2
PENCELL = {}
r = p_first
for fmt, label in (("half", "Half PPR"), ("ppr", "Full PPR")):
    cells = []
    for src in ("Winks", "Norris", "FantasyPros"):
        size = D[fmt]["sizes"][src]
        put(f"A{r}", f"{label} — {src}", 9)
        c = put(f"B{r}", size + 15, 9, True, color="0000FF", fill="FFFF00"); c.border = box
        put(f"C{r}", f"list is {size} deep", 9, color="777777", italic=True)
        cells.append(f"Legend!$B${r}")
        r += 1
    PENCELL[fmt] = tuple(cells)

put(f"A{r+1}", "Blue on yellow are hard-coded inputs you can change. Black is a formula.", 8, color="777777", italic=True)
put(f"A{r+2}", "Your no-kicker league: the extra flex makes RB/WR depth matter more than the K column. Filter by Pos and work past RB60 / WR72.", 9, color="555555")

lg.column_dimensions["A"].width = 30
lg.column_dimensions["B"].width = 60
lg.column_dimensions["C"].width = 30
for i in range(4, len(HDRS) + 2):
    lg.column_dimensions[get_column_letter(i)].width = 9

CW, CN, CF = COL["Winks"], COL["Norris"], COL["FantasyPros"]
flag_note = (f"A shaded Range means the sources disagree by {FLAG} or more."
             + ("  ECR-ADP: green falls to you, red is a room reach." if HAS_ADP else ""))

for fmt, tab in (("half", "Half PPR"), ("ppr", "Full PPR")):
    ws = wb.create_sheet(tab); ws.sheet_view.showGridLines = False
    rows = D[fmt]["rows"]
    ws["A1"] = f"{tab} — Winks + Norris + FantasyPros consensus"
    ws["A1"].font = Font(name=ARIAL, size=13, bold=True)
    ws["A2"] = "Filter with the header arrows. Set Drafted to dim a row. " + flag_note
    ws["A2"].font = Font(name=ARIAL, size=9, color="666666")

    for i, h in enumerate(HDRS):
        c = ws.cell(row=4, column=1+i, value=h)
        c.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HDR)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = box
    ws.row_dimensions[4].height = 20

    pW, pN, pF = PENCELL[fmt]
    for idx, p in enumerate(rows):
        r = 5 + idx
        rk = p["ranks"]
        avg = (f'=(IF({CW}{r}="",{pW},{CW}{r})*{WCELL["Winks"]}'
               f'+IF({CN}{r}="",{pN},{CN}{r})*{WCELL["Norris"]}'
               f'+IF({CF}{r}="",{pF},{CF}{r})*{WCELL["FantasyPros"]})'
               f'/({WCELL["Winks"]}+{WCELL["Norris"]}+{WCELL["FantasyPros"]})')
        rng = f'=IF(COUNT({CW}{r}:{CF}{r})<2,"",MAX({CW}{r}:{CF}{r})-MIN({CW}{r}:{CF}{r}))'
        vals = {"Rank": idx+1, "Tier": p["tier"], "PosTier": p["ptier"], "Player": p["name"],
                "Pos": p["pos"], "PosRank": p["posrank"], "Team": p["team"], "Avg": avg,
                "Winks": rk.get("Winks"), "Norris": rk.get("Norris"),
                "FantasyPros": rk.get("FantasyPros"), "Range": rng,
                "ECR-ADP": p.get("delta"), "Trend": trend_str(p),
                "Drafted": None, "Notes": None}
        for ci, h in enumerate(HDRS):
            c = ws.cell(row=r, column=1+ci, value=vals[h])
            c.font = Font(name=ARIAL, size=9); c.border = box
            if h not in ("Player", "Notes"): c.alignment = Alignment(horizontal="center")
            if h == "Avg": c.number_format = "0.0"
            if h == "Pos":
                c.fill = PatternFill("solid", fgColor=POS_FILL.get(p["pos"], "FFFFFF"))
                c.font = Font(name=ARIAL, size=9, bold=True)
            if h in ("Drafted", "Notes"): c.fill = PatternFill("solid", fgColor="FFF9DB")

    last = 4 + len(rows)
    lastcol = get_column_letter(len(HDRS))
    ws.auto_filter.ref = f"A4:{lastcol}{last}"
    ws.freeze_panes = "A5"
    for i, h in enumerate(HDRS):
        ws.column_dimensions[get_column_letter(i+1)].width = BASE_WIDTH[h]

    dv = DataValidation(type="list", formula1='"ME,TAKEN"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f'{COL["Drafted"]}5:{COL["Drafted"]}{last}')

    rng_all = f"A5:{lastcol}{last}"
    dcol = COL["Drafted"]
    ws.conditional_formatting.add(rng_all, FormulaRule(
        formula=[f'${dcol}5="TAKEN"'], fill=PatternFill("solid", fgColor="EDEDED"),
        font=Font(name=ARIAL, size=9, color="AAAAAA", strike=True)))
    ws.conditional_formatting.add(rng_all, FormulaRule(
        formula=[f'${dcol}5="ME"'], fill=PatternFill("solid", fgColor="DCF5E4")))
    ws.conditional_formatting.add(f'{COL["Range"]}5:{COL["Range"]}{last}', CellIsRule(
        operator="greaterThanOrEqual", formula=[str(FLAG)],
        fill=PatternFill("solid", fgColor="FFE9A8"), font=Font(name=ARIAL, size=9, bold=True)))
    if HAS_ADP:
        acol = COL["ECR-ADP"]
        ws.conditional_formatting.add(f"{acol}5:{acol}{last}", CellIsRule(
            operator="lessThanOrEqual", formula=["-5"],
            fill=PatternFill("solid", fgColor="DCF5E4"), font=Font(name=ARIAL, size=9, bold=True)))
        ws.conditional_formatting.add(f"{acol}5:{acol}{last}", CellIsRule(
            operator="greaterThanOrEqual", formula=["5"],
            fill=PatternFill("solid", fgColor="F8DADA"), font=Font(name=ARIAL, size=9, bold=True)))

wb.save(Path(os.environ.get("FF_OUTPUT_DIR", ROOT / "output")) / "draft-cheat-sheet.xlsx")
print("saved")
