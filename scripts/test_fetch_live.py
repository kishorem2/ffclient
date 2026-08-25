"""End-to-end offline test of the LIVE path: synthetic FantasyPros + Sleeper
API responses -> fetch_live -> full pipeline -> audit, all in a sandbox.
No network, no real cache, no real budget, and data/ is never touched.

Run:  python3 scripts/test_fetch_live.py
"""
import json, os, re, shutil, subprocess, sys, tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sandbox = Path(tempfile.mkdtemp(prefix="fetchlive-demo-"))
sb_data, sb_out = sandbox / "data", sandbox / "output"
sb_data.mkdir(); sb_out.mkdir()
for f in ("winks_half.txt", "winks_full.txt", "norris_half.txt", "norris_full.txt"):
    shutil.copy(ROOT / "data" / f, sb_data / f)

ENV = {**os.environ,
       "FP_CACHE_DIR": str(sandbox / "cache"), "FP_STATE_FILE": str(sandbox / "state.json"),
       "FP_DAILY_BUDGET": "10", "FANTASYPROS_API_KEY": "demo-key",
       "FF_DATA_DIR": str(sb_data), "FF_OUTPUT_DIR": str(sb_out)}
os.environ.update(ENV)

sys.path.insert(0, str(Path(__file__).resolve().parent))
import fpapi, fetch_live
from matchkey import mkey

# ---- synthetic FantasyPros: real static lists as API JSON, deterministic
# rank_std, and an ADP ordering perturbed by a fixed pattern
def load_rows(name):
    rows = []
    for line in open(ROOT / "data" / name):
        line = line.strip()
        if line:
            nm, pos, team = line.split("|")
            rows.append((nm, pos, team))
    return rows

def fp_json(rows, adp=False):
    players = []
    order = list(range(len(rows)))
    if adp:  # the "room" reaches for some, lets others fall
        for i in range(0, len(order) - 7, 7):
            order[i], order[i + 6] = order[i + 6], order[i]
    for rank, idx in enumerate(order, 1):
        nm, pos, team = rows[idx]
        players.append({"player_name": nm, "player_position_id": pos,
                        "player_team_id": team, "rank_ecr": rank,
                        "rank_min": max(1, rank - 5), "rank_max": rank + 5,
                        "rank_ave": float(rank), "rank_std": round(2 + (idx * 7 % 60) / 2, 1),
                        "pos_rank": f"{pos}{rank}", "player_bye_week": (idx % 14) + 4,
                        "player_owned_avg": 99.0})
    return {"total_experts": 111, "last_updated": "Aug 25 2026 (synthetic)", "players": players}

HALF, PPR = load_rows("fp_half.txt"), load_rows("fp_ppr.txt")

# ---- synthetic Sleeper: player map + trending; includes the two trap cases
sleeper_players = {
    "1001": {"full_name": "Jahmyr Gibbs", "position": "RB", "team": "DET"},
    "1002": {"full_name": "Bijan Robinson", "position": "RB", "team": "ATL"},
    # suffixless spelling + JAX-style team code: must still decorate Travis
    # Etienne Jr. (unique once the team disambiguates from Trevor/CAR)
    "1003": {"full_name": "Travis Etienne", "position": "RB", "team": "NO"},
    # Brian Robinson Jr. spelled without the Jr.: strips to the same key as
    # Bijan (both RB/ATL), so the fallback must refuse to guess
    "1004": {"full_name": "Brian Robinson", "position": "RB", "team": "ATL"},
    "1005": {"full_name": "Puka Nacua", "position": "WR", "team": "LA"},  # LA -> LAR
    "HOU": {"position": "DEF", "team": "HOU"},  # Sleeper keys defenses by team code
}
trending_add = [{"player_id": "1002", "count": 40310}, {"player_id": "1003", "count": 9917},
                {"player_id": "1004", "count": 5432}, {"player_id": "1005", "count": 20144},
                {"player_id": "HOU", "count": 3001}]
trending_drop = [{"player_id": "1001", "count": 777}]

def stub(url, params, headers):
    if "consensus-rankings" in url:
        rows = HALF if params.get("scoring") == "HALF" else PPR
        return 200, json.dumps(fp_json(rows, adp=params.get("type") == "ADP"))
    if url.endswith("/v1/players/nfl"):
        return 200, json.dumps(sleeper_players)
    if "trending/add" in url:
        return 200, json.dumps(trending_add)
    if "trending/drop" in url:
        return 200, json.dumps(trending_drop)
    raise AssertionError(f"unexpected url {url}")

fails = 0
def check(label, cond):
    global fails
    print(("  PASS  " if cond else "  FAIL  ") + label)
    if not cond:
        fails += 1

print("== fetch_live against synthetic APIs ==")
fpapi._http_get = stub
fetch_live.main([])

state = json.load(open(ENV["FP_STATE_FILE"]))
check("exactly 4 FP requests (verify + ppr + 2x adp; half was cached by verify)",
      state["count"] == 4)
meta = json.load(open(sb_data / "fp_meta.json"))
check("fp_half.txt written with full row count",
      len((sb_data / "fp_half.txt").read_text().splitlines()) == len(HALF))
bijan = meta["half"]["players"][mkey("Bijan Robinson", "RB", "ATL")]
check("meta carries std + adp + delta ingredients", bijan["std"] is not None and "adp" in bijan)
check("trending normalized LA->LAR onto Puka's mkey",
      mkey("Puka Nacua", "WR", "LAR") in meta["trending"]["add"])
check("DST trending keyed on team", "DST|HOU" in meta["trending"]["add"])

print("\n== full pipeline on the synthetic data ==")
for script in ("agg_three.py", "fix_names.py", "build_four.py", "build_xlsx3.py"):
    r = subprocess.run([sys.executable, str(ROOT / "scripts" / script)], env=ENV,
                       capture_output=True, text=True)
    if r.returncode != 0:
        print(r.stdout, r.stderr); check(f"{script} ran", False); break
else:
    check("pipeline ran clean", True)

r = subprocess.run([sys.executable, str(ROOT / "scripts" / "audit.py")], env=ENV,
                   capture_output=True, text=True)
check("audit ends TOTAL PROBLEMS: 0", "TOTAL PROBLEMS: 0" in r.stdout)

d = json.load(open(sb_data / "consensus.json"))
for fmt in ("half", "ppr"):
    rb = [x for x in d[fmt]["rows"] if x["pos"] == "RB"]
    check(f"Bijan Robinson is RB2 in {fmt}", rb[1]["name"] == "Bijan Robinson")

rows = {r["name"]: r for r in d["half"]["rows"]}
check("std threshold computed", "std" in d["flag"])
check("delta (ECR-ADP) attached to rows",
      any(r.get("delta") not in (None, 0) for r in d["half"]["rows"]))
check("Bijan/suffixless-Brian collision dropped on the Sleeper side (neither decorated)",
      rows["Bijan Robinson"].get("trend_add") is None
      and rows["Brian Robinson Jr."].get("trend_add") is None)
check("suffixless 'Travis Etienne' decorated Travis Etienne Jr. via unique fallback",
      rows["Travis Etienne Jr."].get("trend_add") == 9917)
check("Puka decorated via exact key after LA->LAR normalization",
      rows["Puka Nacua"].get("trend_add") == 20144)
check("HOU defense decorated via team-only DST key",
      next(r for r in d["half"]["rows"] if r["pos"] == "DST" and r["team"] == "HOU").get("trend_add") == 3001)
check("Gibbs carries the drop count", rows["Jahmyr Gibbs"].get("trend_drop") == 777)

from pdfminer.high_level import extract_text
txt = extract_text(str(sb_out / "rankings-half-ppr.pdf"))
check("rankings PDF grew the ±ADP column", "±ADP" in txt)
check("rankings PDF footer explains value vs reach", "value" in txt and "reach" in txt)
cheat = extract_text(str(sb_out / "cheat-sheet-half-ppr.pdf"))
check("cheat sheet footnotes the trending marker", "Sleeper adds" in cheat)

from openpyxl import load_workbook
ws = load_workbook(sb_out / "draft-cheat-sheet.xlsx")["Half PPR"]
hdrs = [c.value for c in ws[4]]
check("workbook has ECR-ADP and Trend columns", "ECR-ADP" in hdrs and "Trend" in hdrs)
di = hdrs.index("Drafted") + 1
dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
from openpyxl.utils import get_column_letter
check("Drafted dropdown follows the moved column",
      any(s.startswith(get_column_letter(di)) for s in dvs))

shutil.rmtree(sandbox)
print(f"\n{'ALL CHECKS PASSED' if fails == 0 else f'{fails} CHECKS FAILED'}")
sys.exit(1 if fails else 0)
